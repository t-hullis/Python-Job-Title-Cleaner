from openpyxl import Workbook, load_workbook
""" Work Book """
wb = load_workbook("Titles.xlsx")
ws = wb.active

# Define the list of strings to check for
strings_to_check = ['Manager']


# Initialize a variable to keep track of the number of matching cells
matching_cells = 0

# Iterate through the rows in column A
for row in ws['A']:
    # Check if the cell value is in the list of strings
    if row.value in strings_to_check:
        matching_cells += 1
        print(row.value)
        print(row.coordinate)

# Print the number of matching cells
print(f"{matching_cells} cells in column A contain a string from the list")