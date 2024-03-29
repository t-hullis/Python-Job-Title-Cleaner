from openpyxl import Workbook, load_workbook
"""Work Book"""
wb = load_workbook("Titles.xlsx")
ws = wb.active
# Initialize a variable to keep track of the number of matching cells
managers= 0
persons = 0

# both departments and personas dictoanries can both be easily edited
departments = {
    "Management": ["Management","CEO","CTO","General Manager", "Founders","Owner", "Director"],
    "Finance":["Finance"],
    "Operations": ["Operations"],
    "IT":[ "I.T","IT"],
    "Product": ["Product"],
    "Support": ["Support"],
    "HR": ["HR"],
    "Sales": ["Sales"],
    "Marketing": ["Marketing",]
    }

personas = {
    "Founder": ["Founder", "Co-Founder"],
    "Owner":["Owner"],
    "CSuite": ["CSuite", "CEO","Chief Executive Office", "CTO", "Cheif Technology Officer", "CFO","Cheif Financial Officer" "COO", "Chief Operating Officer",
     "CGO", "Chief Analytics Officer", "CAO", "Chief Marketing Officer", "CMO", "Chief Data Officer", "CDO"],
    "Managing Director":[ "Managing Director"],
    "Director": ["Director"],
    "Engineer": ["Engineer", "Engineering"],
    "VP": ["VP", "Vice President"],
    "Accounting": ["Accounting","Controller", "accounting", "accountant"],
    "Manager": ["Manager"],
    "Finance": ["Finance"],
    "Legal": ["Legal"]
    }


# Finds department key words in database
for key, values in departments.items():
    # Iterate through the rows in column A
    for value in values: 
        
        for row in ws['A']:
            # Define the list of strings to check for
            strings_to_check = value    
            count = row.value.count(strings_to_check)
            managers = managers + count
            if count == 1:
                # row.coordinate data type is string
                print(row.coordinate)
                print(row.value)
                ws[f"D{row.coordinate[1:]}"] = key
            
# Finds personas key words in database 
for key, values in personas.items():
    # Iterate through the rows in column A
    for value in values: 
    # Iterate through the rows in column A
        for row in ws['A']:
            # Define the list of strings to check for
            strings_to_check = value   
            count = row.value.count(strings_to_check)
            persons = persons + count
            if count==1:
                print(f"persona..{row.coordinate}")
                print(f"persona..{row.value}")
                ws[f"C{row.coordinate[1:]}"] = key

# This for loop creates the cleaned title for column b in the excel sheet
# Iterate through all rows in column C and D, starting from the second row
for row in range(2, ws.max_row + 1):
    # Get the value in cell C and D
    value_c = ws.cell(row=row, column=3).value
    if type(value_c) != type("string"):
        value_c = "_"
    value_d = ws.cell(row=row, column=4).value
    if type(value_d) != type("string"):
        value_d = "_"
    # set the value in column B 
    if value_c != value_d:
        ws.cell(row=row, column=2).value = value_d + " " + value_c 
    else:
        ws.cell(row=row, column=2).value = value_c 


# Print the number of matching cells
print(f"{managers} Manager Titles")
print(f"{persons} Personas")

wb.save('Titles.xlsx')
